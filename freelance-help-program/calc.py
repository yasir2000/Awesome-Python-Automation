import os
from dotenv import load_dotenv
import openpyxl
from datetime import datetime
from tkinter import *
from tkinter import filedialog, messagebox

# Load environment variables from .env file
load_dotenv()

class FileHandler:
    """Handles file operations including loading Excel files."""
    
    @staticmethod
    def load_workbook(file_path):
        return openpyxl.load_workbook(file_path)

class CalculationStrategy:
    """Interface for calculation strategies."""
    
    def calculate(self, file_path, pay_per_hour):
        raise NotImplementedError("This method should be overridden.")

class PaymentCalculation(CalculationStrategy):
    """Concrete strategy for calculating payment based on hours worked."""
    
    def calculate(self, file_path, pay_per_hour):
        wb = FileHandler.load_workbook(file_path)
        sheet = wb.active
        total_price = 0

        for row in range(2, sheet.max_row + 1):
            cell_start = sheet.cell(row, 1)
            cell_finish = sheet.cell(row, 2)

            start_time = datetime.combine(datetime.now().date(), cell_start.value)
            finish_time = datetime.combine(datetime.now().date(), cell_finish.value)

            total_time = finish_time - start_time
            total_time_cell = sheet.cell(row, 3)
            total_time_cell.value = total_time.total_seconds() / 3600

            price = pay_per_hour * (total_time.total_seconds() / 3600)
            price_cell = sheet.cell(row, 4)
            price_cell.value = price
            price_cell.number_format = '0.00'

            total_price += price

        total_price_cell = sheet.cell(sheet.max_row + 1, 4)
        total_price_cell.value = total_price
        total_price_cell.number_format = '0.00'

        modified_file_path = f"modified_{file_path}"
        wb.save(modified_file_path)
        return total_price

class CalculatorApp:
    """Main application for calculating payments."""
    
    def __init__(self, root):
        self.root = root
        self.strategy = PaymentCalculation()
        self.create_widgets()
    
    def create_widgets(self):
        Label(self.root, text="Let's calculate the payment and hours", padx=15, pady=15,
              font="SegoeUI 14", bg="palegreen1", fg="red").grid(row=1, column=1, pady=10, padx=5, columnspan=3)

        Label(self.root, text="XL File Path:", font="SegoeUI 12", bg="PaleGreen1").grid(row=2, column=1, pady=10, padx=5)
        self.file_entry = Entry(self.root, width=30)
        self.file_entry.grid(row=2, column=2, pady=10, padx=5)
        
        Button(self.root, text="Browse", font="SegoeUI 12", bg="PaleGreen1", command=self.browse).grid(row=2, column=3, pady=10, padx=5)

        Label(self.root, text="Pay per Hour:", font="SegoeUI 12", bg="PaleGreen1").grid(row=3, column=1, pady=10, padx=5)
        self.pay_entry = Entry(self.root, width=30)
        self.pay_entry.grid(row=3, column=2, pady=10, padx=5)

        Button(self.root, text="Calculate", font="SegoeUI 12", bg="PaleGreen1", command=self.calculate_button_clicked).grid(row=4, column=1, pady=10, padx=5, columnspan=3)

    def browse(self):
        file_path = filedialog.askopenfilename(title="Select XL File")
        if file_path.endswith(('.xlsx', '.xls')):
            self.file_entry.delete(0, END)
            self.file_entry.insert(0, file_path)
        else:
            messagebox.showerror("Error", "Invalid file format")

    def calculate_button_clicked(self):
        file_path = self.file_entry.get()
        pay_per_hour = self.pay_entry.get()
        if not file_path or not pay_per_hour:
            messagebox.showerror("Error", "Please fill all fields.")
            return
        
        try:
            pay_per_hour = float(pay_per_hour)
            total_payment = self.strategy.calculate(file_path, pay_per_hour)
            messagebox.showinfo("Calculation Complete", f"Total Payment: {total_payment:.2f} Shekels")
        except ValueError:
            messagebox.showerror("Error", "Invalid pay per hour value.")

def main():
    root = Tk()
    root.geometry("520x280")
    root.resizable(False, False)
    root.title("Freelance Help Software")
    root.config(background="PaleGreen1")
    
    app = CalculatorApp(root)
    root.mainloop()

if __name__ == '__main__':
    main()
